from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from fake_useragent import UserAgent
from openpyxl import *
import time
import chromedriver_autoinstaller

chromedriver_autoinstaller.install()  

#건드릴 필요 없음!
user_agent = "Mozilla/5.0 (Linux; Android 9; SM-G975F) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.83 Mobile Safari/537.36" #
options = webdriver.ChromeOptions()
options.add_argument("headless")
options.add_argument('user-agent=' + user_agent)
caps = DesiredCapabilities().CHROME
caps["pageLoadStrategy"] = "none"
driver = webdriver.Chrome(options=options)

revise_list=[]

#크롤링 함수
def MakeChart(financeCode): 
    try:
        codeUrl = "https://dart.fss.or.kr/dsab007/main.do"
        driver.get(codeUrl)
        driver.implicitly_wait(10)

        Select(driver.find_element(By.XPATH,'//*[@id="option"]')).select_by_value('corp')#검색 카테고리(옵션)
        driver.find_element(By.XPATH,'//*[@id="textCrpNm"]').send_keys(financeCode)#종목코드 입력
        driver.find_element(By.XPATH,'//*[@id="li_01"]').click()#정기공시 선택
        driver.find_element(By.XPATH,'//*[@id="publicTypeDetail_A001"]').click()#사업보고서 선택
        driver.find_element(By.XPATH,'//*[@id="searchForm"]/div[2]/div[2]/a[1]').send_keys(Keys.ENTER)#검색
        try:
            target_td=driver.find_element(By.XPATH,'//*[@id="tbody"]/tr/td[3]')#사업보고서 링크 있는 td태그 선택
        except:
            target_td=driver.find_element(By.XPATH,'//*[@id="tbody"]/tr[1]/td[3]')#사업보고서 링크 있는 td태그 선택(사업보고서 링크 2개 이상인 경우 최근 업로드 항목 선택)
            revise_list.append(financeCode)#직접 확인 필요한 항목 리스트에 저장
        target_a=target_td.find_element(By.TAG_NAME,"a")#td태그 하위 a태그 선택
        target_link=target_a.get_attribute("href")#href 링크 정보

        return target_link
       
    except Exception as e:
        print("오류: 해당 종목의 사업보고서가 존재하지 않거나 크롤링 과정에서 문제 발생함("+financeCode+")")
        print(e)#Error message전문 출력
        revise_list.append(financeCode)#직접 확인 필요한 항목 리스트에 저장
        return("#error")
#엑셀 작업 함수
def Report(reading_cell,writing_cell,start_num,end_num,file_path):

    start = time.time() #작업 시간 측정(시작시간)
    wb=Workbook()
    wb=load_workbook(file_path,data_only=True)
    ws=wb['Sheet1'] #작업시트 정보

    get_cells=ws[reading_cell+str(start_num) : reading_cell+str(end_num)]

    for row in get_cells:
        for cell in row:

            target_cell=writing_cell+str(start_num)
            code=cell.value

            print(target_cell+":"+code)

            ws[target_cell].value=MakeChart(code)
            start_num+=1

            wb.save(file_path) 

    print(revise_list)
    end = time.time() #작업 시간 측정(종료시간)
    print(f"{end - start:.5f} sec") #작업 시간 측정(총 경과시간)


my_xlsx='/Users/sunwoo/Desktop/Code/DartReportsCrawler/codeList.xlsx' #엑셀 파일 위치(정확히 넣어야함)
Report("B","C",2,1344,my_xlsx)#종목코드 읽어올 열/링크 입력할 열/시작 번호/끝 번호/엑셀파일 경로 입력

"""
[실행방법]
1. 컴퓨터 네트워크 설정에서 프록시 우회할 도메인 주소를 '183.89.115.39' 로 변경. 호스트는 별도로 설정하지 않아도 되지만 바꿔야 한다면 local로 두면 됨.
2. my_xlsx 변수에 작업할 엑셀파일의 절대경로 입력
3. 메인 함수인 Report의 매개변수에 해당하는 항목들 입력
4. 실행

*주의사항:
1. 초기 실행 시 ModuleNotFoundError: No module named '모듈이름' 과 같은 에러 뜰 수 있음
    해당 경우에는 'pip install 모듈이름' 명령어를 터미널에 직접 입력해서 다운받으면 되고, 확실하게 해결하기 위해선 에러메세지를 구글링해 필요한 모듈을 찾아서 설치하는 것을 권장
    문제 있을 시 개발자와 컨택하시기 바랍니다. 
2. my_xlsx 변수, Report의 매개변수를 제외한 나머지 항목은 임의로 수정하지 말 것.
3. 프록시는 태국 도메인으로 우회하고 있으나 필요 시 타 국가, 타 도메인으로 수정해도 무관
4. 크롬 드라이버와 엑셀 파일은 본 py와 같은 폴더 안에 둘 것. 

© 2023 Sunwoo Jung <sunwoo1137@gmail.com>
"""